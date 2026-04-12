#!/usr/bin/env python3
"""
Excel知識ベース → JSON変換スクリプト
仮説思考支援ツール_内部知識ベース_部位別拡張版.xlsx を読み込み、
src/data/knowledgeBase.json へ変換する。
"""
import json
import sys
import os
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl が必要です: pip3 install openpyxl")
    sys.exit(1)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', '仮説思考支援ツール_内部知識ベース_部位別拡張版.xlsx')
OUTPUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'src', 'data', 'knowledgeBase.json')


def read_sheet(wb, name):
    """シートをdict配列として読み込み（ヘッダー行を自動検出）"""
    if name not in wb.sheetnames:
        print(f"  [SKIP] シート '{name}' が見つかりません")
        return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []

    # ヘッダー行を自動検出: 'RuleID' or 'SourceKey' or 'RegionID' を含む行
    header_idx = 0
    for idx, row in enumerate(rows):
        if not row or len(row) == 0:
            continue
        first_cell = str(row[0]).strip() if row[0] else ''
        if first_cell in ('RuleID', 'SourceKey', 'RegionID', 'ProblemID', '対象群',
                          'RuleID ', 'col_0'):
            header_idx = idx
            break
        # 2セル以上が非空で、かつ1文字以上の短い識別子っぽい列が複数あればヘッダー
        non_empty = [c for c in row if c is not None and str(c).strip()]
        if len(non_empty) >= 3 and all(len(str(c)) < 100 for c in non_empty):
            # Check it's not a title row (single merged cell with long text)
            if row[1] is not None and str(row[1]).strip():
                header_idx = idx
                break

    headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[header_idx])]
    # Remove 'None' headers
    headers = [h if h != 'None' else f'col_{i}' for i, h in enumerate(headers)]

    result = []
    for row in rows[header_idx + 1:]:
        if not any(c is not None for c in row):
            continue
        d = {}
        for i, val in enumerate(row):
            if i < len(headers):
                d[headers[i]] = str(val).strip() if val is not None else ''
        result.append(d)
    return result


def safe_id(prefix, idx, label=''):
    """安全なID生成"""
    return f"{prefix}-{idx:03d}"


# ─── 参考文献変換 ───
def convert_sources(rows, prefix):
    sources = []
    for row in rows:
        sk = row.get('SourceKey', '') or row.get('Source_key', '')
        if not sk:
            continue
        sources.append({
            'source_key': sk,
            'domain': row.get('領域', ''),
            'title': (row.get('タイトル', '') or row.get('テーマ', '')
                      or row.get('主要知見要約', '')),
            'year': row.get('年', ''),
            'study_design': row.get('研究デザイン', ''),
            'evidence_level': (row.get('EvidenceLevel', '')
                               or row.get('エビデンスレベル', '')),
            'population': row.get('対象', '') or row.get('対象集団', ''),
            'practical_use': (row.get('実務上の使いどころ', '')
                              or row.get('主要知見要約', '')),
            'url': row.get('URL', '') or row.get('PMID/DOI/URL', ''),
            'note': row.get('備考', '') or row.get('注意点', ''),
        })
    return sources


# ─── 観察→テスト対応変換 ───
def convert_obs_test_links(rows, layer, items, links):
    for i, row in enumerate(rows):
        rule_id = row.get('RuleID', '') or row.get('ルールID', '') or f'{layer}-ot-{i+1:03d}'
        obs_text = (row.get('頻出観察所見', '') or row.get('観察所見', '')
                    or row.get('問診から想定される主要テーマ', '') or row.get('観察所見・場面', ''))
        test_text = (row.get('推奨テスト候補', '') or row.get('テスト候補', '')
                     or row.get('優先する観察', '') or row.get('優先観察候補', ''))
        rationale = row.get('何を確かめるか', '') or row.get('狙い', '') or row.get('観察の目的', '')
        condition = row.get('条件変更/比較', '') or row.get('比較条件', '')
        priority = row.get('優先度', 'medium')
        source_keys = row.get('SourceKey', '') or row.get('Source_key', '')
        region = row.get('RegionID', '') or row.get('部位', '')
        region_label = row.get('部位', '')
        movement = row.get('動作テーマID', '') or row.get('動作', '')
        population = row.get('対象群ID', '') or row.get('対象群', '')
        condition_theme = (row.get('症候群/サブテーマ', '') or row.get('課題テーマ', '')
                           or row.get('サブテーマ', '') or row.get('主訴/場面', ''))

        obs_id = f'{layer}-obs-{i+1:03d}'
        test_id = f'{layer}-test-{i+1:03d}'

        items.append({
            'id': obs_id,
            'item_type': 'observation',
            'title': obs_text,
            'content': obs_text,
            'body_region': region,
            'body_region_label': region_label,
            'movement_theme': movement,
            'target_population': population,
            'condition_theme': condition_theme,
            'priority_base': normalize_priority(priority),
            'source_keys': source_keys,
            'layer': layer,
        })

        items.append({
            'id': test_id,
            'item_type': 'test',
            'title': test_text,
            'content': rationale,
            'body_region': region,
            'body_region_label': region_label,
            'movement_theme': movement,
            'target_population': population,
            'condition_theme': condition_theme,
            'priority_base': normalize_priority(priority),
            'source_keys': source_keys,
            'layer': layer,
        })

        links.append({
            'id': f'{layer}-otl-{i+1:03d}',
            'from_id': obs_id,
            'to_id': test_id,
            'link_type': 'observation_to_test',
            'rationale': rationale,
            'condition_note': condition,
            'priority_weight': priority_to_weight(priority),
            'source_keys': source_keys,
        })


# ─── テスト→評価対応変換 ───
def convert_test_eval_links(rows, layer, items, links):
    for i, row in enumerate(rows):
        test_text = (row.get('主要テスト所見', '') or row.get('テスト所見パターン', '')
                     or row.get('テスト結果パターン', '') or row.get('テスト所見・場面', '')
                     or row.get('テスト所見', ''))
        eval_text = (row.get('評価候補', '') or row.get('評価解釈候補', '')
                     or row.get('評価解釈', ''))
        eval_type = row.get('種別', '') or row.get('位置づけ', '')
        rationale = (row.get('なぜそう解釈するか', '') or row.get('解釈の根拠', '')
                     or row.get('解釈根拠', '') or row.get('解釈の軸', ''))
        next_check = (row.get('次に確認したいこと', '') or row.get('次の確認', '')
                      or row.get('次に見たいこと', ''))
        priority = row.get('優先度', 'medium')
        source_keys = row.get('SourceKey', '') or row.get('Source_key', '')
        region = row.get('RegionID', '') or row.get('部位', '')
        region_label = row.get('部位', '')
        movement = row.get('動作テーマID', '') or row.get('動作', '') or row.get('MotionID', '') or row.get('動作テーマ', '')
        population = row.get('対象群ID', '') or row.get('対象群', '')
        condition_theme = (row.get('症候群/サブテーマ', '') or row.get('課題テーマ', '')
                           or row.get('サブテーマ', '') or row.get('主訴/場面', ''))

        test_id = f'{layer}-testfind-{i+1:03d}'
        eval_id = f'{layer}-eval-{i+1:03d}'

        items.append({
            'id': test_id,
            'item_type': 'test',
            'title': test_text,
            'content': test_text,
            'body_region': region,
            'body_region_label': region_label,
            'movement_theme': movement,
            'target_population': population,
            'condition_theme': condition_theme,
            'priority_base': normalize_priority(priority),
            'source_keys': source_keys,
            'layer': layer,
        })

        items.append({
            'id': eval_id,
            'item_type': 'evaluation',
            'title': eval_text,
            'content': rationale,
            'body_region': region,
            'body_region_label': region_label,
            'movement_theme': movement,
            'target_population': population,
            'condition_theme': condition_theme,
            'priority_base': normalize_priority(priority),
            'source_keys': source_keys,
            'layer': layer,
            'eval_type': eval_type,
            'next_check': next_check,
        })

        links.append({
            'id': f'{layer}-tel-{i+1:03d}',
            'from_id': test_id,
            'to_id': eval_id,
            'link_type': 'test_to_evaluation',
            'rationale': rationale,
            'condition_note': next_check,
            'priority_weight': priority_to_weight(priority),
            'source_keys': source_keys,
        })


# ─── 評価→介入対応変換 ───
def convert_eval_intervention_links(rows, layer, items, links):
    for i, row in enumerate(rows):
        eval_text = row.get('評価候補', '') or row.get('評価解釈', '') or row.get('評価カテゴリ', '')
        iv_category = row.get('介入カテゴリ', '') or row.get('カテゴリ', '')
        iv_text = row.get('介入候補', '') or row.get('介入メニュー候補', '')
        iv_intent = row.get('介入意図', '') or row.get('意図', '')
        reeval = row.get('主な再評価項目', '') or row.get('再評価', '')
        condition = row.get('適用条件', '') or row.get('備考', '')
        source_keys = row.get('SourceKey', '') or row.get('Source_key', '')
        region = row.get('RegionID', '') or row.get('部位', '')
        region_label = row.get('部位', '')
        movement = row.get('動作テーマID', '') or row.get('動作', '')
        population = row.get('対象群ID', '') or row.get('対象群', '')
        condition_theme = (row.get('症候群/サブテーマ', '') or row.get('課題テーマ', '')
                           or row.get('サブテーマ', '') or row.get('主訴/場面', ''))

        eval_id = f'{layer}-evalfind-{i+1:03d}'
        iv_id = f'{layer}-iv-{i+1:03d}'
        reeval_id = f'{layer}-reeval-{i+1:03d}'

        items.append({
            'id': eval_id,
            'item_type': 'evaluation',
            'title': eval_text,
            'content': eval_text,
            'body_region': region,
            'body_region_label': region_label,
            'movement_theme': movement,
            'target_population': population,
            'condition_theme': condition_theme,
            'source_keys': source_keys,
            'layer': layer,
        })

        items.append({
            'id': iv_id,
            'item_type': 'intervention',
            'title': iv_text,
            'content': iv_intent,
            'body_region': region,
            'body_region_label': region_label,
            'movement_theme': movement,
            'target_population': population,
            'condition_theme': condition_theme,
            'source_keys': source_keys,
            'layer': layer,
            'intervention_category': iv_category,
        })

        links.append({
            'id': f'{layer}-eil-{i+1:03d}',
            'from_id': eval_id,
            'to_id': iv_id,
            'link_type': 'evaluation_to_intervention',
            'rationale': iv_intent,
            'condition_note': condition,
            'priority_weight': 60,
            'source_keys': source_keys,
        })

        # 再評価アイテム
        if reeval:
            items.append({
                'id': reeval_id,
                'item_type': 'reevaluation',
                'title': reeval,
                'content': reeval,
                'body_region': region,
                'body_region_label': region_label,
                'movement_theme': movement,
                'target_population': population,
                'condition_theme': condition_theme,
                'source_keys': source_keys,
                'layer': layer,
            })
            links.append({
                'id': f'{layer}-irl-{i+1:03d}',
                'from_id': iv_id,
                'to_id': reeval_id,
                'link_type': 'intervention_to_reevaluation',
                'rationale': f'{iv_text} の効果を確認',
                'priority_weight': 50,
                'source_keys': source_keys,
            })


# ─── 問診→観察対応変換 ───
def convert_intake_obs_links(rows, layer, items, links):
    for i, row in enumerate(rows):
        intake_text = (row.get('問診から想定される主要テーマ', '') or row.get('問診テーマ', '')
                       or row.get('問診トリガー', '') or row.get('主訴/目標カテゴリ', ''))
        obs_text = (row.get('優先する観察', '') or row.get('観察候補', '')
                    or row.get('優先観察候補', '') or row.get('優先観察カテゴリ', ''))
        rationale = row.get('狙い', '') or row.get('根拠', '') or row.get('観察の目的', '')
        source_keys = row.get('SourceKey', '') or row.get('Source_key', '')
        condition_theme = (row.get('課題テーマ', '') or row.get('サブテーマ', '')
                           or row.get('主訴/場面', '') or row.get('主訴/目標カテゴリ', ''))

        intake_id = f'{layer}-intake-{i+1:03d}'
        obs_id = f'{layer}-intobs-{i+1:03d}'

        items.append({
            'id': intake_id,
            'item_type': 'intake_question',
            'title': intake_text,
            'content': intake_text,
            'condition_theme': condition_theme,
            'source_keys': source_keys,
            'layer': layer,
        })

        items.append({
            'id': obs_id,
            'item_type': 'observation',
            'title': obs_text,
            'content': rationale,
            'condition_theme': condition_theme,
            'source_keys': source_keys,
            'layer': layer,
        })

        links.append({
            'id': f'{layer}-iol-{i+1:03d}',
            'from_id': intake_id,
            'to_id': obs_id,
            'link_type': 'intake_to_observation',
            'rationale': rationale,
            'priority_weight': 60,
            'source_keys': source_keys,
        })


# ─── 優先マトリクス変換 ───
def convert_priority_matrix(rows, rule_type):
    rules = []
    for i, row in enumerate(rows):
        keys = list(row.keys())
        rule_key = row.get(keys[0], '') if keys else ''
        note = row.get('使い方メモ', '') or row.get('備考', '')
        source_keys = row.get('SourceKey', '')

        # 各動作/症候群列を処理
        for key in keys[1:]:
            if key in ('使い方メモ', '備考', 'SourceKey', 'RegionID', '部位'):
                continue
            val = row.get(key, '')
            if not val:
                continue

            if rule_type == 'region_syndrome':
                # 部位×症候群マトリクス
                rules.append({
                    'rule_type': rule_type,
                    'rule_key': row.get('RegionID', '') or row.get(keys[0], ''),
                    'condition_key': row.get('症候群/サブテーマ', '') or key,
                    'boost_observations': row.get('AI初期表示で上位に出す観察軸', ''),
                    'boost_tests': row.get('AI初期表示で上位に出すテスト', ''),
                    'boost_interventions': row.get('AI初期表示で上位に出す介入', ''),
                    'boost_reevaluations': row.get('主な再評価', ''),
                    'weight_adjustment': 0,
                    'source_keys': source_keys,
                    'note': note,
                })
                break  # 1行1ルール
            else:
                # 対象×動作マトリクス
                try:
                    weight = int(float(val))
                except (ValueError, TypeError):
                    weight = 1
                rules.append({
                    'rule_type': rule_type,
                    'rule_key': rule_key,
                    'condition_key': key,
                    'weight_adjustment': weight * 10,
                    'source_keys': source_keys,
                    'note': note,
                })

    return rules


def normalize_priority(p):
    p = str(p).strip().lower()
    if p in ('高', 'high', '3'):
        return 'high'
    if p in ('低', 'low', '1'):
        return 'low'
    return 'medium'


def priority_to_weight(p):
    p = normalize_priority(p)
    return {'high': 80, 'medium': 60, 'low': 40}.get(p, 60)


def main():
    print(f"読み込み: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    print(f"シート数: {len(wb.sheetnames)}")

    items = []
    links = []
    sources = []
    priority_rules = []

    # ─── 基本層 ───
    print("\n[1/5] 基本層...")
    intake_obs = read_sheet(wb, '問診_観察対応')
    convert_intake_obs_links(intake_obs, 'base', items, links)

    obs_test = read_sheet(wb, '観察_テスト対応')
    convert_obs_test_links(obs_test, 'base', items, links)

    test_eval = read_sheet(wb, 'テスト_評価対応')
    convert_test_eval_links(test_eval, 'base', items, links)

    eval_iv = read_sheet(wb, '評価_介入対応')
    convert_eval_intervention_links(eval_iv, 'base', items, links)

    base_sources = read_sheet(wb, 'エビデンス台帳')
    sources.extend(convert_sources(base_sources, 'base'))

    # ─── エビデンス拡張層 ───
    print("[2/5] エビデンス拡張層...")
    eb_intake_obs = read_sheet(wb, 'EB_問診_観察対応')
    convert_intake_obs_links(eb_intake_obs, 'evidence', items, links)

    eb_obs_test = read_sheet(wb, 'EB_観察_テスト対応')
    convert_obs_test_links(eb_obs_test, 'evidence', items, links)

    eb_test_eval = read_sheet(wb, 'EB_テスト_評価対応')
    convert_test_eval_links(eb_test_eval, 'evidence', items, links)

    eb_eval_iv = read_sheet(wb, 'EB_評価_介入対応')
    convert_eval_intervention_links(eb_eval_iv, 'evidence', items, links)

    eb_sources = read_sheet(wb, 'EB_参考文献')
    sources.extend(convert_sources(eb_sources, 'evidence'))

    # ─── 動作別層 ───
    print("[3/5] 動作別層...")
    mv_obs_test = read_sheet(wb, '動作別_観察_テスト対応')
    convert_obs_test_links(mv_obs_test, 'movement', items, links)

    mv_test_eval = read_sheet(wb, '動作別_テスト_評価対応')
    convert_test_eval_links(mv_test_eval, 'movement', items, links)

    mv_eval_iv = read_sheet(wb, '動作別_評価_介入対応')
    convert_eval_intervention_links(mv_eval_iv, 'movement', items, links)

    mv_sources = read_sheet(wb, '動作別_参考文献')
    sources.extend(convert_sources(mv_sources, 'movement'))

    # ─── 対象別層 ───
    print("[4/5] 対象別層...")
    tg_obs_test = read_sheet(wb, '対象別_観察_テスト対応')
    convert_obs_test_links(tg_obs_test, 'population', items, links)

    tg_test_eval = read_sheet(wb, '対象別_テスト_評価対応')
    convert_test_eval_links(tg_test_eval, 'population', items, links)

    tg_eval_iv = read_sheet(wb, '対象別_評価_介入対応')
    convert_eval_intervention_links(tg_eval_iv, 'population', items, links)

    tg_sources = read_sheet(wb, '対象別_参考文献')
    sources.extend(convert_sources(tg_sources, 'population'))

    # 対象×動作マトリクス
    tg_matrix = read_sheet(wb, '対象×動作_優先マトリクス')
    priority_rules.extend(convert_priority_matrix(tg_matrix, 'population_movement'))

    # ─── 部位別層 ───
    print("[5/5] 部位別層...")
    rg_obs_test = read_sheet(wb, '部位別_観察_テスト対応')
    convert_obs_test_links(rg_obs_test, 'region', items, links)

    rg_test_eval = read_sheet(wb, '部位別_テスト_評価対応')
    convert_test_eval_links(rg_test_eval, 'region', items, links)

    rg_eval_iv = read_sheet(wb, '部位別_評価_介入対応')
    convert_eval_intervention_links(rg_eval_iv, 'region', items, links)

    rg_sources = read_sheet(wb, '部位別_参考文献')
    sources.extend(convert_sources(rg_sources, 'region'))

    # 部位×症候群マトリクス
    rg_matrix = read_sheet(wb, '部位×症候群_優先マトリクス')
    priority_rules.extend(convert_priority_matrix(rg_matrix, 'region_syndrome'))

    # ─── 出力 ───
    kb = {
        'version': '1.0',
        'generated_at': datetime.now().isoformat(),
        'stats': {
            'sources': len(sources),
            'items': len(items),
            'links': len(links),
            'priority_rules': len(priority_rules),
            'items_by_type': {},
            'items_by_layer': {},
            'links_by_type': {},
        },
        'sources': sources,
        'items': items,
        'links': links,
        'priority_rules': priority_rules,
    }

    # Stats
    for item in items:
        t = item['item_type']
        kb['stats']['items_by_type'][t] = kb['stats']['items_by_type'].get(t, 0) + 1
        l = item.get('layer', 'unknown')
        kb['stats']['items_by_layer'][l] = kb['stats']['items_by_layer'].get(l, 0) + 1
    for link in links:
        t = link['link_type']
        kb['stats']['links_by_type'][t] = kb['stats']['links_by_type'].get(t, 0) + 1

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(kb, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 変換完了: {OUTPUT_PATH}")
    print(f"  参考文献: {len(sources)}件")
    print(f"  アイテム: {len(items)}件")
    for t, c in kb['stats']['items_by_type'].items():
        print(f"    {t}: {c}")
    print(f"  リンク: {len(links)}件")
    for t, c in kb['stats']['links_by_type'].items():
        print(f"    {t}: {c}")
    print(f"  優先ルール: {len(priority_rules)}件")


if __name__ == '__main__':
    main()
